from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Paleta F1-08
COR_TITULO = "1F4E79"
COR_CABECALHO = "2E75B6"
COR_TOTAL = "70AD47"
COR_BORDA = "D9D9D9"

_BORDA_FINA = Border(
    left=Side(style="thin", color=COR_BORDA),
    right=Side(style="thin", color=COR_BORDA),
    top=Side(style="thin", color=COR_BORDA),
    bottom=Side(style="thin", color=COR_BORDA),
)


def _preencher_celula(
    celula,
    *,
    valor: Any,
    fundo: str | None = None,
    negrito: bool = False,
    branco: bool = False,
    alinhamento: str = "left",
    numero: bool = False,
) -> None:
    celula.value = valor
    celula.font = Font(bold=negrito, color="FFFFFF" if branco else "000000")
    if fundo:
        celula.fill = PatternFill("solid", fgColor=fundo)
    celula.border = _BORDA_FINA
    horizontal = "right" if numero else alinhamento
    celula.alignment = Alignment(horizontal=horizontal, vertical="center", wrap_text=True)


def _larguras_colunas(ws, pesos: list[float]) -> None:
    for idx, peso in enumerate(pesos, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = 12 * peso


def _aba_quantitativos(ws, dados: dict[str, Any]) -> None:
    colunas = dados["colunas"]
    linhas = dados["linhas"]
    casas = dados.get("casas_decimais", 4)
    total_geral = dados.get("total_geral")

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(colunas))
    titulo = ws.cell(row=1, column=1, value="Quantitativos")
    _preencher_celula(titulo, valor="Quantitativos", fundo=COR_TITULO, negrito=True, branco=True)
    titulo.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    for col_idx, nome in enumerate(colunas, start=1):
        cel = ws.cell(row=2, column=col_idx, value=nome)
        _preencher_celula(cel, valor=nome, fundo=COR_CABECALHO, negrito=True, branco=True, alinhamento="center")

    for row_idx, linha in enumerate(linhas, start=3):
        for col_idx, valor in enumerate(linha, start=1):
            numero = col_idx > 1 and isinstance(valor, (int, float))
            cel = ws.cell(row=row_idx, column=col_idx)
            _preencher_celula(cel, valor=valor, numero=numero)
            if numero:
                cel.number_format = f"#,##0.{'0' * casas}"

    if total_geral is not None and linhas:
        row_total = 3 + len(linhas)
        ws.cell(row=row_total, column=1, value="TOTAL GERAL")
        _preencher_celula(
            ws.cell(row=row_total, column=1),
            valor="TOTAL GERAL",
            fundo=COR_TOTAL,
            negrito=True,
            branco=True,
        )
        for col_idx in range(2, len(colunas) + 1):
            soma = sum(
                (linha[col_idx - 1] or 0)
                for linha in linhas
                if col_idx - 1 < len(linha) and isinstance(linha[col_idx - 1], (int, float))
            )
            valor = soma if soma else total_geral if col_idx == 2 else None
            cel = ws.cell(row=row_total, column=col_idx, value=valor)
            _preencher_celula(cel, valor=valor, fundo=COR_TOTAL, negrito=True, branco=True, numero=True)
            if isinstance(valor, (int, float)):
                cel.number_format = f"#,##0.{'0' * casas}"

    pesos = [2.0] + [1.5] + [1.0] * max(0, len(colunas) - 2)
    _larguras_colunas(ws, pesos[: len(colunas)])
    ws.freeze_panes = "A3"


def _aba_detalhamento(ws, dados: dict[str, Any]) -> None:
    cabecalho = ["Fonte", "Estilo", "Área (ha)", "Feições", "Geometrias corrigidas"]
    for col_idx, nome in enumerate(cabecalho, start=1):
        cel = ws.cell(row=1, column=col_idx, value=nome)
        _preencher_celula(cel, valor=nome, fundo=COR_CABECALHO, negrito=True, branco=True, alinhamento="center")

    casas = dados.get("casas_decimais", 4)
    for row_idx, item in enumerate(dados.get("detalhe") or [], start=2):
        valores = [
            item.get("fonte"),
            item.get("estilo"),
            item.get("area_ha"),
            item.get("feicoes"),
            item.get("geometrias_corrigidas"),
        ]
        for col_idx, valor in enumerate(valores, start=1):
            cel = ws.cell(row=row_idx, column=col_idx, value=valor)
            _preencher_celula(cel, valor=valor, numero=col_idx == 3 and isinstance(valor, (int, float)))
            if col_idx == 3 and isinstance(valor, (int, float)):
                cel.number_format = f"#,##0.{'0' * casas}"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{max(1, 1 + len(dados.get('detalhe') or []))}"
    _larguras_colunas(ws, [2.0, 1.5, 1.2, 1.0, 1.2])


def _aba_avisos(ws, avisos: list[str]) -> None:
    ws.cell(row=1, column=1, value="Aviso")
    _preencher_celula(ws.cell(row=1, column=1), valor="Aviso", fundo=COR_CABECALHO, negrito=True, branco=True)
    for row_idx, texto in enumerate(avisos or ["Nenhum aviso."], start=2):
        cel = ws.cell(row=row_idx, column=1, value=texto)
        _preencher_celula(cel, valor=texto)
        cel.font = Font(italic=True, color="666666", size=9)
    ws.column_dimensions["A"].width = 80
    ws.auto_filter.ref = f"A1:A{max(1, len(avisos) + 1)}"


def _aba_fontes(ws, dados: dict[str, Any]) -> None:
    cabecalho = ["Fonte", "Estilo", "Arquivo", "CRS"]
    for col_idx, nome in enumerate(cabecalho, start=1):
        cel = ws.cell(row=1, column=col_idx, value=nome)
        _preencher_celula(cel, valor=nome, fundo=COR_CABECALHO, negrito=True, branco=True, alinhamento="center")

    for row_idx, item in enumerate(dados.get("detalhe") or [], start=2):
        valores = [
            item.get("fonte"),
            item.get("estilo"),
            item.get("arquivo", "—"),
            item.get("crs", "—"),
        ]
        for col_idx, valor in enumerate(valores, start=1):
            _preencher_celula(ws.cell(row=row_idx, column=col_idx), valor=valor)

    _larguras_colunas(ws, [1.5, 1.5, 2.5, 1.5])


def _aba_conferencia(ws, conferencia: dict[str, Any], *, casas: int = 4) -> None:
    cabecalho = [
        "Classe",
        "Declarado no recibo (ha)",
        "Calculado (ha)",
        "Diferença (ha)",
        "Diferença (%)",
        "OK",
    ]
    for col_idx, nome in enumerate(cabecalho, start=1):
        cel = ws.cell(row=1, column=col_idx, value=nome)
        _preencher_celula(cel, valor=nome, fundo=COR_CABECALHO, negrito=True, branco=True, alinhamento="center")

    linhas = conferencia.get("linhas") or []
    if not linhas:
        msg = (
            "Sem recibo do CAR no workspace — abre a pasta com o PDF do recibo "
            "para preencher esta aba."
            if not conferencia.get("tem_recibo")
            else "Sem classes para conferir."
        )
        cel = ws.cell(row=2, column=1, value=msg)
        _preencher_celula(cel, valor=msg)
        cel.font = Font(italic=True, color="666666", size=9)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    else:
        for row_idx, item in enumerate(linhas, start=2):
            valores = [
                item.get("classe"),
                item.get("declarado_ha"),
                item.get("calculado_ha"),
                item.get("diferenca_ha"),
                item.get("diferenca_pct"),
                "sim" if item.get("ok") else "não",
            ]
            for col_idx, valor in enumerate(valores, start=1):
                cel = ws.cell(row=row_idx, column=col_idx, value=valor)
                numero = col_idx in (2, 3, 4) and isinstance(valor, (int, float))
                pct = col_idx == 5 and isinstance(valor, (int, float))
                _preencher_celula(cel, valor=valor, numero=numero or pct)
                if numero:
                    cel.number_format = f"#,##0.{'0' * casas}"
                if pct:
                    cel.number_format = "0.00%"
                if col_idx == 6 and not item.get("ok"):
                    cel.font = Font(bold=True, color="C00000")

    ws.freeze_panes = "A2"
    _larguras_colunas(ws, [2.5, 1.8, 1.5, 1.5, 1.3, 0.8])


def exportar_xlsx(
    dados: dict[str, Any],
    destino: Path,
    *,
    fontes_detalhe: list[dict[str, Any]] | None = None,
    recibo: dict[str, Any] | None = None,
) -> Path:
    """Gera planilha de quantitativos (F1-08) a partir do resultado de calcular()."""
    from mapasfacil_nucleo.quantitativos.conferencia import montar_conferencia

    destino.parent.mkdir(parents=True, exist_ok=True)

    if fontes_detalhe:
        detalhe = list(dados.get("detalhe") or [])
        por_fonte = {d.get("fonte"): d for d in detalhe}
        for extra in fontes_detalhe:
            if extra.get("fonte") in por_fonte:
                por_fonte[extra["fonte"]].update(extra)
        dados = {**dados, "detalhe": list(por_fonte.values())}

    conferencia = dados.get("conferencia")
    if conferencia is None:
        conferencia = montar_conferencia(dados, recibo)

    wb = Workbook()
    ws_q = wb.active
    ws_q.title = "Quantitativos"
    _aba_quantitativos(ws_q, dados)

    ws_d = wb.create_sheet("Detalhamento")
    _aba_detalhamento(ws_d, dados)

    ws_c = wb.create_sheet("Conferência")
    _aba_conferencia(ws_c, conferencia, casas=int(dados.get("casas_decimais") or 4))

    avisos = list(dados.get("avisos") or [])
    avisos.extend(conferencia.get("avisos") or [])
    ws_a = wb.create_sheet("Avisos")
    _aba_avisos(ws_a, avisos)

    ws_f = wb.create_sheet("Fontes")
    _aba_fontes(ws_f, dados)

    wb.save(destino)
    return destino
