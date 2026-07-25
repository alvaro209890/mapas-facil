from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from mapasfacil_nucleo.fsguard import WorkspaceGuard
from mapasfacil_nucleo.motores.gerar import gerar_mapa
from mapasfacil_nucleo.quantitativos.calcular import calcular as calcular_quantitativos
from mapasfacil_nucleo.quantitativos.xlsx import exportar_xlsx
from tests.helpers_fixtures import montar_workspace_minimo


def test_exportar_xlsx_abas_e_estilo(tmp_path: Path) -> None:
    montar_workspace_minimo(tmp_path)
    guard = WorkspaceGuard(tmp_path)
    fontes_idx = {
        "ATP": "dados/ATP.shp",
        "AVN": "dados/AVN.shp",
        "AUAS": "dados/AUAS.shp",
    }
    mapspec = {
        "imovel": {"nome": "Fazenda Teste"},
        "camadas": [
            {"fonte": "local.ATP", "estilo": "perimetro_imovel"},
            {"fonte": "local.AVN", "estilo": "avn"},
            {"fonte": "local.AUAS", "estilo": "auas"},
        ],
        "tabela": {"total_geral": True, "casas_decimais": 4},
    }
    dados = calcular_quantitativos(mapspec, guard=guard, fontes_idx=fontes_idx)
    destino = tmp_path / "Mapas" / "teste_Quantitativos.xlsx"
    exportar_xlsx(dados, destino)

    assert destino.exists()
    wb = load_workbook(destino)
    assert wb.sheetnames == ["Quantitativos", "Detalhamento", "Avisos", "Fontes"]
    ws = wb["Quantitativos"]
    assert ws["A1"].value == "Quantitativos"
    assert ws["A1"].fill.fgColor.rgb.endswith("1F4E79")
    assert ws["A3"].value == "Fazenda Teste"
    assert ws["B3"].value == 100.0


def test_gerar_mapa_com_xlsx(tmp_path: Path, repo_root: Path) -> None:
    import copy
    import json

    montar_workspace_minimo(tmp_path)
    caminho = repo_root / "shared/fixtures/mapspecs/dinamica_2026_canonico.json"
    mapspec = copy.deepcopy(json.loads(caminho.read_text(encoding="utf-8")))
    mapspec["camadas"] = [c for c in mapspec["camadas"] if c["fonte"].startswith("local.")]
    mapspec["saidas"] = ["pdf", "xlsx"]
    mapspec["saida"] = {
        "pasta": "Mapas",
        "nome_base": "Dinamica_xlsx",
        "materializar_camadas_em": "SHP",
    }

    fontes_idx = {
        "ATP": "dados/ATP.shp",
        "AVN": "dados/AVN.shp",
        "AUAS": "dados/AUAS.shp",
    }
    guard = WorkspaceGuard(tmp_path)
    resultado = gerar_mapa(mapspec, guard, fontes_idx)
    xlsx = tmp_path / resultado["xlsx"]
    assert xlsx.exists()
    assert xlsx.name == "Dinamica_xlsx_Quantitativos.xlsx"
